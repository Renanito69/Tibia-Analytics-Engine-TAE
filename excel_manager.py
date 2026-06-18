from openpyxl import load_workbook, Workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
import os
from datetime import datetime
from config import ARQUIVO_EXCEL

COLUNAS = ["Data", "exp", "loot", "profit", "tempo_excel", "exp_hora"]

COLUNAS_LOOT_DETALHADO = [
    "hunt_id", "item", "quantidade", "npc", "valor_unidade", "valor_total"
]

COLUNAS_RESUMO = ["Balance", "Total Farmado", "Tempo jogado", "Xp Total"]

COLUNAS_FORMULAS = [
    "=SUM(D:D)",
    "=SUM(C:C)",
    "=SUM(E:E)",
    "=SUM(B:B)"
]

def salvar_dados(metricas):
    metricas["Data"] = datetime.now().strftime("%d/%m/%y")
    if not os.path.exists(ARQUIVO_EXCEL):
        wb = Workbook()
        ws = wb.active
        ws.title = "Resumo"

        ws.append(COLUNAS)

    else:
        wb = load_workbook(ARQUIVO_EXCEL)

        if "Resumo" not in wb.sheetnames:
            ws = wb.create_sheet("Resumo")
            ws.append(COLUNAS)
        else:
            ws = wb["Resumo"]

    nova_linha = [metricas.get(col, None) for col in COLUNAS]
    ws.append(nova_linha)

    if ws["H1"].value is None:
        for i, valor in enumerate(COLUNAS_RESUMO):
            ws.cell(row=1, column=8 + i, value=valor)

        for i, formula in enumerate(COLUNAS_FORMULAS):
            ws.cell(row=2, column=8 + i, value=formula)
        
        tabela_kpi = ws.tables.get("TabelaResumoKPIs")

        if tabela_kpi:
            tabela_kpi.ref = "H1:K2"
        else:
            tabela_kpi = Table(
                displayName="TabelaResumoKPIs",
                ref="H1:K2"
            )

            tabela_kpi.tableStyleInfo = TableStyleInfo(
                name="TableStyleLight8",
                showRowStripes=True
            )

            ws.add_table(tabela_kpi)

    ultima_linha = ws.max_row

    tabela = ws.tables.get("TabelaResumo")

    if tabela:
        tabela.ref = f"A1:F{ultima_linha}"
    else:
        tabela = Table(
            displayName="TabelaResumo",
            ref=f"A1:F{ultima_linha}"
        )

        tabela.tableStyleInfo = TableStyleInfo(
            name="TableStyleLight8",
            showRowStripes=True
        )

        ws.add_table(tabela)

    wb.save(ARQUIVO_EXCEL)
    print("[OK] Dados salvos na aba Resumo")


# =========================
# FUNÇÃO LOOT
# =========================
def salvar_loot(loot_lista, hunt_id):

    # =========================
    # GARANTIR ARQUIVO
    # =========================
    if not os.path.exists(ARQUIVO_EXCEL):
        wb = Workbook()
        wb.save(ARQUIVO_EXCEL)

    wb = load_workbook(ARQUIVO_EXCEL)

    # =========================
    # CRIAR OU OBTER ABA LOOT
    # =========================
    if "Loot" not in wb.sheetnames:
        ws = wb.create_sheet("Loot")
        ws.append(COLUNAS_LOOT_DETALHADO)
    else:
        ws = wb["Loot"]

    # =========================
    # KPI TOTAL (I1:I2)
    # =========================
    if ws["I1"].value is None:
        ws["I1"] = "Total"
        ws["I2"] = "=SUM(F:F)"

    # Criar/garantir tabela do KPI
    tabela_kpi = ws.tables.get("TabelaLootKPI")

    if tabela_kpi:
        tabela_kpi.ref = "I1:I2"
    else:
        tabela_kpi = Table(
            displayName="TabelaLootKPI",
            ref="I1:I2"
        )

        tabela_kpi.tableStyleInfo = TableStyleInfo(
            name="TableStyleLight8",
            showRowStripes=True
        )

        ws.add_table(tabela_kpi)

    # =========================
    # INSERIR LOOT
    # =========================
    for item in loot_lista:
        ws.append([
            hunt_id,
            item["item"],
            item["quantidade"],
            item["npc"],
            item["valor_unidade"],
            item["valor_total"]
        ])

    # =========================
    # TABELA LOOT (dados)
    # =========================
    ultima_linha = ws.max_row

    tabela = ws.tables.get("TabelaLoot")

    if tabela:
        tabela.ref = f"A1:F{ultima_linha}"
    else:
        tabela = Table(
            displayName="TabelaLoot",
            ref=f"A1:F{ultima_linha}"
        )

        tabela.tableStyleInfo = TableStyleInfo(
            name="TableStyleLight8",
            showRowStripes=True
        )

        ws.add_table(tabela)

    wb.save(ARQUIVO_EXCEL)
    print("[OK] Loot salvo com sucesso")